import pandas as pd

from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

class RawDataTemplate:
    def __init__(self):
        self.template_path = Path(__file__).parent.parent.parent / "upload_templates" / "raw_data_template.xlsx"

        self.meta_data = {
            "model": {"format": str,
                      "impossible": {"type": "length", "min": 1, "max": 30}},
            "latitude_deg": {"format": float,
                             "impossible": {"type": "range", "min": -90, "max": 90}},
            "longitude_deg": {"format": float,
                              "impossible": {"type": "range", "min": -180, "max": 180}},
            "hub_height_m": {"format": float,
                             "impossible": {"type": "range", "min": 0, "max": 400},
                             "plausible": {"type": "range", "min": 50, "max": 200}},
            "rotor_diameter_m": {"format": float,
                                 "impossible": {"type": "range", "min": 0, "max": 400},
                                 "plausible": {"type": "range", "min": 40, "max": 280}},
            "installation_year": {"format": int,
                                  "impossible": {"type": "range", "min": 1980, "max": datetime.now().year}},
            "onshore_offshore": {"format": str,
                                 "impossible": {"type": "enum", "values": ["onshore", "offshore"]}},
        }

        self.mandatory = {
            "timestamp": {"format": "%Y-%m-%d %H:%M:%S",
                          "impossible": {"type": "range", "min": 1980, "max": datetime.now().year}},
            "wind_speed_ms": {"format": float,
                              "impossible": {"type": "range", "min": 0.0, "max": 80.0},
                              "plausible": {"type": "range", "min": 0.0, "max": 35.0}},
            "power_kw": {"format": float,
                         "impossible": {"type": "range", "min": -500.0, "max": 30_000.0},
                         "plausible": {"type": "range", "min": -100.0, "max": 25_000.0}},
            "wind_direction_deg": {"format": float,
                                   "impossible": {"type": "range", "min": 0, "max": 360}},
            "status_code": {"format": str,
                            "impossible": {"type": "length", "min": 1, "max": 200}},
        }

        self.optional = {
            "wind_speed_avg_10min_ms": {"format": float,
                                        "impossible": {"type": "range", "min": 0.0, "max": 80.0},
                                        "plausible": {"type": "range", "min": 0.0, "max": 35.0}},
            "wind_speed_max_ms": {"format": float,
                                  "impossible": {"type": "range", "min": 0.0, "max": 80.0},
                                  "plausible": {"type": "range", "min": 0.0, "max": 50.0}},
            "wind_speed_min_ms": {"format": float,
                                  "impossible": {"type": "range", "min": 0.0, "max": 80.0},
                                  "plausible": {"type": "range", "min": 0.0, "max": 35.0}},
            "wind_shear_exp": {"format": float,
                               "impossible": {"type": "range", "min": 0.0, "max": 1.0}},
            "wind_turbulence_intensity_pct": {"format": float,
                                              "impossible": {"type": "range", "min": 0.0, "max": 100.0},
                                              "plausible": {"type": "range", "min": 0.0, "max": 60.0}},
            "air_density_kgm3": {"format": float,
                                 "impossible": {"type": "range", "min": 0.9, "max": 1.4}},
            "air_temperature_celsius": {"format": float,
                                        "impossible": {"type": "range", "min": -50.0, "max": 60.0},
                                        "plausible": {"type": "range", "min": -40.0, "max": 50.0}},
            "air_pressure_hpa": {"format": float,
                                 "impossible": {"type": "range", "min": 870.0, "max": 1084.0}},
            "air_humidity_pct": {"format": float,
                                 "impossible": {"type": "range", "min": 0.0, "max": 100.0}},
            "icing_detected_bool": {"format": bool,
                                    "impossible": {"type": "bool"}},
            "rotor_speed_rpm": {"format": float,
                                "impossible": {"type": "range", "min": 0.0, "max": 25.0},
                                "plausible": {"type": "range", "min": 0.0, "max": 20.0}},
            "rotor_torque_nm": {"format": float,
                                "impossible": {"type": "range", "min": 0.0, "max": 20_000_000.0}},
            "rotor_azimuth_position_deg": {"format": float,
                                           "impossible": {"type": "range", "min": 0.0, "max": 360.0}},
            "rotor_imbalance_pct": {"format": float,
                                    "impossible": {"type": "range", "min": 0.0, "max": 100.0},
                                    "plausible": {"type": "range", "min": 0.0, "max": 20.0}},
            "blade_pitch_angle_1_deg": {"format": float,
                                        "impossible": {"type": "range", "min": -5.0, "max": 95.0}},
            "blade_pitch_angle_2_deg": {"format": float,
                                        "impossible": {"type": "range", "min": -5.0, "max": 95.0}},
            "blade_pitch_angle_3_deg": {"format": float,
                                        "impossible": {"type": "range", "min": -5.0, "max": 95.0}},
            "blade_root_bending_moment_1_knm": {"format": float,
                                                "impossible": {"type": "range", "min": 0.0, "max": 100_000.0}},
            "blade_root_bending_moment_2_knm": {"format": float,
                                                "impossible": {"type": "range", "min": 0.0, "max": 100_000.0}},
            "blade_root_bending_moment_3_knm": {"format": float,
                                                "impossible": {"type": "range", "min": 0.0, "max": 100_000.0}},
            "blade_vibration_1_ms2": {"format": float,
                                      "impossible": {"type": "range", "min": 0.0, "max": 50.0},
                                      "plausible": {"type": "range", "min": 0.0, "max": 20.0}},
            "blade_vibration_2_ms2": {"format": float,
                                      "impossible": {"type": "range", "min": 0.0, "max": 50.0},
                                      "plausible": {"type": "range", "min": 0.0, "max": 20.0}},
            "blade_vibration_3_ms2": {"format": float,
                                      "impossible": {"type": "range", "min": 0.0, "max": 50.0},
                                      "plausible": {"type": "range", "min": 0.0, "max": 20.0}},
            "generator_speed_rpm": {"format": float,
                                    "impossible": {"type": "range", "min": 0.0, "max": 2000.0},
                                    "plausible": {"type": "range", "min": 0.0, "max": 1800.0}},
            "generator_torque_nm": {"format": float,
                                    "impossible": {"type": "range", "min": 0.0, "max": 500_000.0}},
            "generator_power_kw": {"format": float,
                                   "impossible": {"type": "range", "min": -500.0, "max": 30_000.0},
                                   "plausible": {"type": "range", "min": -100.0, "max": 25_000.0}},
            "generator_current_a": {"format": float,
                                    "impossible": {"type": "range", "min": 0.0, "max": 5000.0}},
            "generator_voltage_v": {"format": float,
                                    "impossible": {"type": "range", "min": 0.0, "max": 35_000.0}},
            "generator_frequency_hz": {"format": float,
                                       "impossible": {"type": "range", "min": 45.0, "max": 65.0}},
            "generator_temperature_stator_celsius": {"format": float,
                                                     "impossible": {"type": "range", "min": -30.0, "max": 180.0},
                                                     "plausible": {"type": "range", "min": -20.0, "max": 155.0}},
            "generator_temperature_rotor_celsius": {"format": float,
                                                    "impossible": {"type": "range", "min": -30.0, "max": 180.0},
                                                    "plausible": {"type": "range", "min": -20.0, "max": 155.0}},
            "generator_temperature_bearing_drive_celsius": {"format": float,
                                                            "impossible": {"type": "range", "min": -30.0, "max": 120.0},
                                                            "plausible": {"type": "range", "min": -20.0, "max": 100.0}},
            "generator_temperature_bearing_nondrive_celsius": {"format": float,
                                                               "impossible": {"type": "range", "min": -30.0,
                                                                              "max": 120.0},
                                                               "plausible": {"type": "range", "min": -20.0,
                                                                             "max": 100.0}},
            "generator_vibration_ms2": {"format": float,
                                        "impossible": {"type": "range", "min": 0.0, "max": 50.0},
                                        "plausible": {"type": "range", "min": 0.0, "max": 20.0}},
            "gearbox_temperature_oil_celsius": {"format": float,
                                                "impossible": {"type": "range", "min": -30.0, "max": 120.0},
                                                "plausible": {"type": "range", "min": -20.0, "max": 90.0}},
            "gearbox_temperature_bearing_hs_celsius": {"format": float,
                                                       "impossible": {"type": "range", "min": -30.0, "max": 120.0},
                                                       "plausible": {"type": "range", "min": -20.0, "max": 100.0}},
            "gearbox_oil_pressure_bar": {"format": float,
                                         "impossible": {"type": "range", "min": 0.0, "max": 10.0}},
            "gearbox_oil_level_mm": {"format": float,
                                     "impossible": {"type": "range", "min": 0.0, "max": 1000.0}},
            "gearbox_vibration_ms2": {"format": float,
                                      "impossible": {"type": "range", "min": 0.0, "max": 50.0},
                                      "plausible": {"type": "range", "min": 0.0, "max": 20.0}},
            "nacelle_position_deg": {"format": float,
                                     "impossible": {"type": "range", "min": 0.0, "max": 360.0}},
            "nacelle_yaw_error_deg": {"format": float,
                                      "impossible": {"type": "range", "min": -180.0, "max": 180.0},
                                      "plausible": {"type": "range", "min": -30.0, "max": 30.0}},
            "nacelle_yaw_rate_degs": {"format": float,
                                      "impossible": {"type": "range", "min": -5.0, "max": 5.0}},
            "nacelle_temperature_celsius": {"format": float,
                                            "impossible": {"type": "range", "min": -30.0, "max": 80.0},
                                            "plausible": {"type": "range", "min": -20.0, "max": 65.0}},
            "nacelle_vibration_x_ms2": {"format": float,
                                        "impossible": {"type": "range", "min": 0.0, "max": 50.0},
                                        "plausible": {"type": "range", "min": 0.0, "max": 20.0}},
            "nacelle_vibration_y_ms2": {"format": float,
                                        "impossible": {"type": "range", "min": 0.0, "max": 50.0},
                                        "plausible": {"type": "range", "min": 0.0, "max": 20.0}},
            "nacelle_vibration_z_ms2": {"format": float,
                                        "impossible": {"type": "range", "min": 0.0, "max": 50.0},
                                        "plausible": {"type": "range", "min": 0.0, "max": 20.0}},
            "tower_top_acceleration_x_ms2": {"format": float,
                                             "impossible": {"type": "range", "min": -10.0, "max": 10.0},
                                             "plausible": {"type": "range", "min": -5.0, "max": 5.0}},
            "tower_top_acceleration_y_ms2": {"format": float,
                                             "impossible": {"type": "range", "min": -10.0, "max": 10.0},
                                             "plausible": {"type": "range", "min": -5.0, "max": 5.0}},
            "tower_fore_aft_bending_moment_knm": {"format": float,
                                                  "impossible": {"type": "range", "min": 0.0, "max": 500_000.0}},
            "tower_side_side_bending_moment_knm": {"format": float,
                                                   "impossible": {"type": "range", "min": 0.0, "max": 500_000.0}},
            "active_power_kw": {"format": float,
                                "impossible": {"type": "range", "min": -500.0, "max": 30_000.0},
                                "plausible": {"type": "range", "min": -100.0, "max": 25_000.0}},
            "reactive_power_kvar": {"format": float,
                                    "impossible": {"type": "range", "min": -10_000.0, "max": 10_000.0}},
            "power_factor_cos_phi": {"format": float,
                                     "impossible": {"type": "range", "min": -1.0, "max": 1.0}},
            "energy_produced_kwh": {"format": float,
                                    "impossible": {"type": "range", "min": 0.0, "max": 30_000.0}},
            "energy_produced_total_mwh": {"format": float,
                                          "impossible": {"type": "range", "min": 0.0, "max": 10_000_000.0}},
            "grid_frequency_hz": {"format": float,
                                  "impossible": {"type": "range", "min": 45.0, "max": 65.0},
                                  "plausible": {"type": "range", "min": 49.0, "max": 51.0}},
            "grid_voltage_l1_v": {"format": float,
                                  "impossible": {"type": "range", "min": 0.0, "max": 35_000.0}},
            "grid_voltage_l2_v": {"format": float,
                                  "impossible": {"type": "range", "min": 0.0, "max": 35_000.0}},
            "grid_voltage_l3_v": {"format": float,
                                  "impossible": {"type": "range", "min": 0.0, "max": 35_000.0}},
            "main_bearing_temperature_celsius": {"format": float,
                                                 "impossible": {"type": "range", "min": -30.0, "max": 120.0},
                                                 "plausible": {"type": "range", "min": -20.0, "max": 100.0}},
            "main_bearing_vibration_ms2": {"format": float,
                                           "impossible": {"type": "range", "min": 0.0, "max": 50.0},
                                           "plausible": {"type": "range", "min": 0.0, "max": 20.0}},
            "hydraulic_pressure_bar": {"format": float,
                                       "impossible": {"type": "range", "min": 0.0, "max": 300.0}},
            "hydraulic_oil_temperature_celsius": {"format": float,
                                                  "impossible": {"type": "range", "min": -30.0, "max": 100.0},
                                                  "plausible": {"type": "range", "min": -20.0, "max": 80.0}},
            "cooling_water_temperature_inlet_celsius": {"format": float,
                                                        "impossible": {"type": "range", "min": -10.0, "max": 80.0}},
            "cooling_water_temperature_outlet_celsius": {"format": float,
                                                         "impossible": {"type": "range", "min": -10.0, "max": 80.0}},
            "cooling_water_flow_rate_lmin": {"format": float,
                                             "impossible": {"type": "range", "min": 0.0, "max": 1000.0}},
            "converter_temperature_celsius": {"format": float,
                                              "impossible": {"type": "range", "min": -30.0, "max": 100.0},
                                              "plausible": {"type": "range", "min": -20.0, "max": 80.0}},
            "converter_dc_voltage_v": {"format": float,
                                       "impossible": {"type": "range", "min": 0.0, "max": 1500.0}},
            "converter_igbt_temperature_celsius": {"format": float,
                                                   "impossible": {"type": "range", "min": -30.0, "max": 150.0},
                                                   "plausible": {"type": "range", "min": -20.0, "max": 125.0}},
            "converter_efficiency_pct": {"format": float,
                                         "impossible": {"type": "range", "min": 0.0, "max": 100.0},
                                         "plausible": {"type": "range", "min": 80.0, "max": 100.0}},
            "brake_pressure_bar": {"format": float,
                                   "impossible": {"type": "range", "min": 0.0, "max": 300.0}},
            "brake_temperature_celsius": {"format": float,
                                          "impossible": {"type": "range", "min": -30.0, "max": 300.0},
                                          "plausible": {"type": "range", "min": -20.0, "max": 200.0}},
            "brake_applied_bool": {"format": bool,
                                   "impossible": {"type": "bool"}},
            "emergency_stop_active_bool": {"format": bool,
                                           "impossible": {"type": "bool"}},
            "maintenance_mode_bool": {"format": bool,
                                      "impossible": {"type": "bool"}},
            "availability_bool": {"format": bool,
                                  "impossible": {"type": "bool"}},
            "fault_code": {"format": str,
                           "impossible": {"type": "length", "min": 1, "max": 20}},
            "warning_code": {"format": str,
                             "impossible": {"type": "length", "min": 1, "max": 20}},
            "capacity_factor_pct": {"format": float,
                                    "impossible": {"type": "range", "min": 0.0, "max": 100.0}},
            "performance_ratio_pct": {"format": float,
                                      "impossible": {"type": "range", "min": 0.0, "max": 120.0}},
            "curtailment_power_kw": {"format": float,
                                     "impossible": {"type": "range", "min": 0.0, "max": 30_000.0}},
            "noise_level_db": {"format": float,
                               "impossible": {"type": "range", "min": 0.0, "max": 120.0},
                               "plausible": {"type": "range", "min": 0.0, "max": 80.0}},
            "shadow_flicker_hours": {"format": float,
                                     "impossible": {"type": "range", "min": 0.0, "max": 24.0}},
            "foundation_tilt_deg": {"format": float,
                                    "impossible": {"type": "range", "min": -5.0, "max": 5.0},
                                    "plausible": {"type": "range", "min": -1.0, "max": 1.0}},
            "control_system_uptime_s": {"format": float,
                                        "impossible": {"type": "range", "min": 0.0, "max": 31_536_000.0}},
        }

        # Alle Felder zusammenführen
        self.data = {**self.mandatory, **self.optional}

        # DataFrames aus den Keys erzeugen
        self.meta_df = pd.DataFrame(columns=list(self.meta_data))
        self.data_df = pd.DataFrame(columns=list(self.data.keys()))

    def generate_xlsx(self):
        wb = Workbook()

        ws_meta = wb.active
        ws_meta.title = "meta_data"

        ws_data = wb.create_sheet("raw_data")

        meta_fields = list(self.meta_data.keys())
        col_width = 0
        for row, value in enumerate(meta_fields, start=1):
            ws_meta[f"A{row}"] = value
            ws_meta[f"A{row}"].font = Font(bold=True)
            width = len(str(value or ""))
            if width > col_width:
                col_width = width

        next_row = len(meta_fields) + 2
        ws_meta[f"A{next_row}"] = "Coordinates will be altered so that the exact location of the turbine will not show."
        ws_meta[f"A{next_row}"].font = Font(bold=True, color="FF0000")

        ws_meta.column_dimensions["A"].width = col_width + 2

        data_fields = list(self.data.keys())
        ws_data.append(data_fields)

        for cell in ws_data[1]:
            cell.font = Font(bold=True)
            col_letter = get_column_letter(cell.column)
            ws_data.column_dimensions[col_letter].width = 25

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        background_color = "f0ebeb"
        
        for col in range(1, 6):
            col_letter = get_column_letter(col)
            for row in range(1, 1001):
                ws_data[f"{col_letter}{row}"].fill = PatternFill(start_color=background_color, end_color=background_color, fill_type="solid")
                ws_data[f"{col_letter}{row}"].border = border

        ws_data.insert_rows(1)
        ws_data.merge_cells("A1:E1")
        ws_data["A1"] = "MANDATORY FIELDS"
        ws_data["A1"].font = Font(bold=True, size=14)
        ws_data["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_data["A1"].fill = PatternFill(start_color=background_color, end_color=background_color, fill_type="solid")
        wb.save(self.template_path)


template = RawDataTemplate()

template.generate_xlsx()
